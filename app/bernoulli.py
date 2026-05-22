import re
import unicodedata
from io import StringIO
from pathlib import Path

import pandas as pd

from app.bernoulli_client import buscar_aluno_bernoulli
from app.settings import PASTA_RUNTIME
from app.time_utils import agora_local


EXTENSOES_PLANILHA = {".csv", ".xlsx", ".xls"}
PASTA_BERNOULLI = PASTA_RUNTIME / "bernoulli"
TOTAL_QUESTOES_BERNOULLI = 90
DIA_BERNOULLI_PRIMEIRO = "1"
DIA_BERNOULLI_SEGUNDO = "2"


def _emitir_progresso(
    progresso_callback,
    percentual: int,
    mensagem: str,
    tipo: str | None = None,
    titulo: str | None = None,
    descricao: str = "",
    resultado: dict | None = None
):
    if not progresso_callback:
        return

    progresso_callback(
        percentual,
        mensagem,
        tipo=tipo,
        titulo=titulo,
        descricao=descricao,
        resultado=resultado
    )


def _normalizar_texto(valor: str) -> str:
    texto = str(valor or "").strip()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = texto.lower()
    texto = re.sub(r"\s+", " ", texto)
    return texto


def _normalizar_ra(valor) -> str:
    texto = str(valor or "").strip()

    if texto.endswith(".0"):
        texto = texto[:-2]

    digitos = "".join(ch for ch in texto if ch.isdigit())
    return digitos


def _normalizar_resposta(valor) -> str:
    texto = str(valor or "").strip().upper()

    if not texto:
        return "*"

    textos_marcacao_dupla = {
        "MULTIPLA_MARCACAO",
        "MULTIPLA MARCACAO",
        "MULTIPLA",
        "DUPLA_MARCACAO",
        "DUPLA MARCACAO",
        "DUPLA",
        "EM_BRANCO",
        "EM BRANCO",
        "BRANCO",
        "NULL",
        "NAN",
        "-"
    }

    if texto in textos_marcacao_dupla:
        return "*"

    letras = re.findall(r"[A-E]", texto)

    if len(letras) == 1 and len(texto) == 1:
        return letras[0]

    if len(set(letras)) != 1 or len(letras) != 1:
        return "*"

    return letras[0]


def _ler_planilha(caminho: Path) -> pd.DataFrame:
    extensao = caminho.suffix.lower()

    def ler_html_disfarcado():
        try:
            conteudo = caminho.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            conteudo = caminho.read_text(encoding="latin-1")

        try:
            tabelas = pd.read_html(StringIO(conteudo), flavor=["bs4"])
        except ImportError as erro:
            raise ValueError(
                "Nao foi possivel ler a tabela HTML exportada pelo sistema porque faltam "
                f"dependencias de parser no ambiente: {erro}"
            ) from erro

        if not tabelas:
            raise ValueError("Nenhuma tabela HTML foi encontrada no arquivo informado.")

        return tabelas[0].fillna("")

    if extensao == ".csv":
        try:
            return pd.read_csv(caminho, dtype=str, encoding="utf-8-sig", sep=None, engine="python")
        except Exception:
            return pd.read_csv(caminho, dtype=str, encoding="latin-1", sep=None, engine="python")

    if extensao == ".xlsx":
        try:
            return pd.read_excel(caminho, dtype=str)
        except ImportError as erro:
            raise ValueError(
                "Nao foi possivel ler o arquivo .xlsx porque a dependencia necessaria nao "
                f"esta instalada: {erro}"
            ) from erro
        except Exception as erro:
            raise ValueError(f"Falha ao ler o arquivo .xlsx: {erro}") from erro

    if extensao == ".xls":
        try:
            return pd.read_excel(caminho, dtype=str, engine="xlrd")
        except ImportError as erro:
            raise ValueError(
                "Nao foi possivel ler o arquivo .xls porque a dependencia xlrd nao esta "
                f"instalada no ambiente atual: {erro}"
            ) from erro
        except Exception as erro:
            cabecalho = caminho.read_bytes()[:256].lstrip()
            if cabecalho.startswith(b"<") or b"<html" in cabecalho.lower() or b"<table" in cabecalho.lower():
                try:
                    return ler_html_disfarcado()
                except Exception as erro_html:
                    raise ValueError(
                        "O arquivo .xls parece ser uma tabela HTML exportada pelo sistema, "
                        f"mas a leitura falhou: {erro_html}"
                    ) from erro_html
            raise ValueError(f"Falha ao ler o arquivo .xls: {erro}") from erro

    raise ValueError(f"Formato não suportado: {caminho.name}")


def _mapa_colunas(df: pd.DataFrame) -> dict:
    return {
        _normalizar_texto(coluna): coluna
        for coluna in df.columns
    }


def _encontrar_coluna(df: pd.DataFrame, candidatos: list[str]) -> str:
    mapa = _mapa_colunas(df)

    for candidato in candidatos:
        coluna = mapa.get(_normalizar_texto(candidato))
        if coluna:
            return coluna

    return ""


def _promover_linha_para_cabecalho(df: pd.DataFrame, indice_linha: int) -> pd.DataFrame:
    if df.empty or indice_linha >= len(df.index):
        return df

    novo_cabecalho = [
        str(valor).strip() if str(valor).strip() else f"coluna_{indice}"
        for indice, valor in enumerate(df.iloc[indice_linha].tolist(), start=1)
    ]

    df_ajustado = df.iloc[indice_linha + 1:].copy()
    df_ajustado.columns = novo_cabecalho
    return df_ajustado.reset_index(drop=True).fillna("")


def _garantir_cabecalho_util(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    candidatos_necessarios = [
        "RA",
        "Registro Acadêmico",
        "Registro Academico",
        "Aluno",
        "Nome",
        "Questao 1",
        "q1",
        "Lingua Estrangeira",
        "Lingua Estrangeira (ING ou ESP)",
    ]

    if any(_encontrar_coluna(df, [candidato]) for candidato in candidatos_necessarios):
        return df

    limite = min(4, len(df.index))

    for indice in range(limite):
        df_candidato = _promover_linha_para_cabecalho(df, indice)

        if (
            _encontrar_coluna(df_candidato, ["RA", "Registro Acadêmico", "Registro Academico"])
            and _encontrar_coluna(df_candidato, ["Aluno", "Nome", "Nome do Aluno"])
        ):
            return df_candidato

    return df


def _detectar_colunas_questoes(df: pd.DataFrame) -> list[tuple[int, str]]:
    questoes = []

    for coluna in df.columns:
        texto = _normalizar_texto(coluna)

        match_questao = re.match(r"questao\s*(\d+)$", texto)
        match_q = re.match(r"q\s*(\d+)$", texto)
        match = match_questao or match_q

        if match:
            questoes.append((int(match.group(1)), coluna))

    questoes.sort(key=lambda item: item[0])
    return questoes


def _mapa_colunas_questoes(df: pd.DataFrame) -> dict[int, str]:
    return {
        numero: coluna
        for numero, coluna in _detectar_colunas_questoes(df)
    }


def _construir_mapa_rmb_por_api(
    serie_ra: pd.Series,
    serie_nome: pd.Series,
    progresso_callback=None
) -> tuple[dict[str, str], dict]:
    mapa_rmb = {}
    consultas = []
    motivos_nao_encontrados = []

    for valor_ra, valor_nome in zip(serie_ra.tolist(), serie_nome.tolist()):
        ra = _normalizar_ra(valor_ra)
        nome = str(valor_nome or "").strip()
        if ra and ra not in mapa_rmb:
            consultas.append((ra, nome))
            mapa_rmb[ra] = ""

    total_consultas = len(consultas)
    ultimo_percentual = None

    for indice, (ra, nome) in enumerate(consultas, start=1):
        resultado = buscar_aluno_bernoulli(ra=ra, nome=nome)

        if not resultado.get("encontrado"):
            motivo = str(resultado.get("motivo", "") or "").strip()
            if motivo and motivo not in motivos_nao_encontrados:
                motivos_nao_encontrados.append(motivo)
        else:
            mapa_rmb[ra] = str(resultado.get("rmb", "") or "").strip()

        if total_consultas:
            percentual = 45 + int((indice / total_consultas) * 35)

            if percentual != ultimo_percentual or indice == total_consultas:
                ultimo_percentual = percentual
                _emitir_progresso(
                    progresso_callback,
                    percentual,
                    f"Consultando RMB na API do Bernoulli ({indice}/{total_consultas})..."
                )

    return mapa_rmb, {
        "consultados": len(consultas),
        "encontrados": sum(1 for valor in mapa_rmb.values() if str(valor).strip()),
        "motivos_nao_encontrados": motivos_nao_encontrados[:3],
    }


def _ajustar_largura_colunas(caminho_saida: Path):
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(caminho_saida)
    worksheet = workbook.active

    for indice, coluna in enumerate(worksheet.iter_cols(1, worksheet.max_column), start=1):
        maior = 0

        for celula in coluna:
            valor = "" if celula.value is None else str(celula.value)
            maior = max(maior, len(valor))

        worksheet.column_dimensions[get_column_letter(indice)].width = min(max(maior + 2, 10), 38)

    workbook.save(caminho_saida)


def _normalizar_dia_bernoulli(dia) -> str:
    valor = str(dia or DIA_BERNOULLI_PRIMEIRO).strip()

    if valor not in {DIA_BERNOULLI_PRIMEIRO, DIA_BERNOULLI_SEGUNDO}:
        raise ValueError("Dia do Simulado Bernoulli inválido. Use 1 ou 2.")

    return valor


def transformar_vetor_keepedu_para_bernoulli(
    caminho_arquivo_keepedu: Path,
    dia: str = DIA_BERNOULLI_PRIMEIRO,
    progresso_callback=None
):
    dia = _normalizar_dia_bernoulli(dia)
    primeiro_numero_questao = 1 if dia == DIA_BERNOULLI_PRIMEIRO else 91
    ultimo_numero_questao = primeiro_numero_questao + TOTAL_QUESTOES_BERNOULLI - 1

    _emitir_progresso(
        progresso_callback,
        10,
        "Lendo a planilha do KeepEdu..."
    )
    df = _garantir_cabecalho_util(_ler_planilha(caminho_arquivo_keepedu).fillna(""))
    _emitir_progresso(
        progresso_callback,
        20,
        "Planilha lida. Validando colunas obrigatorias..."
    )

    coluna_ra = _encontrar_coluna(df, ["RA", "Registro Acadêmico", "Registro Academico"])
    coluna_nome = _encontrar_coluna(df, ["Aluno", "Nome", "Nome do Aluno"])
    coluna_lingua = _encontrar_coluna(
        df,
        [
            "Lingua Estrangeira",
            "Lingua Estrangeira (ING ou ESP)",
            "Língua Estrangeira",
            "Língua Estrangeira (ING ou ESP)"
        ]
    )
    colunas_questoes = _mapa_colunas_questoes(df)

    if not coluna_ra:
        raise ValueError("Não foi possível localizar a coluna RA no arquivo do KeepEdu.")

    if not coluna_nome:
        raise ValueError("Não foi possível localizar a coluna Aluno/Nome no arquivo do KeepEdu.")

    if dia == DIA_BERNOULLI_PRIMEIRO and not coluna_lingua:
        raise ValueError("Não foi possível localizar a coluna de Língua Estrangeira no arquivo do KeepEdu.")

    if not colunas_questoes:
        raise ValueError("Nenhuma coluna de questão foi encontrada no arquivo do KeepEdu.")

    _emitir_progresso(
        progresso_callback,
        30,
        "Estrutura validada. Preparando consulta Bernoulli...",
        tipo="success",
        titulo="Estrutura validada",
        descricao=f"{len(colunas_questoes)} coluna(s) de questao reconhecida(s).",
        resultado={
            "resultado": {
                "linhas": int(len(df)),
                "questoes_encontradas_no_keepedu": int(len(colunas_questoes))
            }
        }
    )
    _emitir_progresso(
        progresso_callback,
        40,
        "Consultando RMB na API do Bernoulli...",
        tipo="info",
        titulo="Consulta Bernoulli iniciada",
        descricao="Buscando os alunos na API do Bernoulli."
    )

    mapa_rmb, diagnostico_rmb = _construir_mapa_rmb_por_api(
        df[coluna_ra],
        df[coluna_nome],
        progresso_callback=progresso_callback
    )

    saida = pd.DataFrame()
    saida["RMB"] = df[coluna_ra].apply(lambda valor: mapa_rmb.get(_normalizar_ra(valor), ""))
    saida["RA"] = df[coluna_ra].apply(_normalizar_ra)
    saida["Nome"] = df[coluna_nome].astype(str).str.strip()
    if dia == DIA_BERNOULLI_SEGUNDO:
        saida["Lingua Estrangeira (ING ou ESP)"] = "N/A"
    else:
        saida["Lingua Estrangeira (ING ou ESP)"] = df[coluna_lingua].astype(str).str.strip()

    for indice, numero in enumerate(range(primeiro_numero_questao, ultimo_numero_questao + 1), start=1):
        coluna_origem = colunas_questoes.get(indice)

        if coluna_origem:
            saida[f"q{numero}"] = df[coluna_origem].apply(_normalizar_resposta)
        else:
            saida[f"q{numero}"] = ""

    total_sem_rmb = int((saida["RMB"].astype(str).str.strip() == "").sum())
    _emitir_progresso(
        progresso_callback,
        82,
        "Consulta Bernoulli concluida. Gerando arquivo final...",
        tipo="success",
        titulo="Consulta Bernoulli concluida",
        descricao=f"{diagnostico_rmb.get('encontrados', 0)} RMB(s) localizado(s) e {total_sem_rmb} pendencia(s) de inscricao.",
        resultado={
            "resultado": {
                "linhas": int(len(saida)),
                "questoes": TOTAL_QUESTOES_BERNOULLI,
                "questoes_encontradas_no_keepedu": int(len(colunas_questoes)),
                "primeira_questao": primeiro_numero_questao,
                "ultima_questao": ultimo_numero_questao,
                "dia": dia,
                "sem_rmb": total_sem_rmb,
                "origem_rmb": "api_bernoulli",
                "rmb_encontrados": int(diagnostico_rmb.get("encontrados", 0)),
            }
        }
    )
    _emitir_progresso(
        progresso_callback,
        90,
        "Montando o arquivo XLSX final..."
    )
    PASTA_BERNOULLI.mkdir(parents=True, exist_ok=True)
    timestamp = agora_local().strftime("%Y%m%d_%H%M%S")
    caminho_saida = PASTA_BERNOULLI / f"vetor_bernoulli_{timestamp}.xlsx"

    saida.to_excel(caminho_saida, index=False)
    _emitir_progresso(
        progresso_callback,
        96,
        "Ajustando o arquivo final para download..."
    )
    _ajustar_largura_colunas(caminho_saida)

    if len(saida) > 0 and total_sem_rmb == len(saida):
        detalhes = diagnostico_rmb.get("motivos_nao_encontrados") or []
        detalhe_texto = f" Detalhes: {' | '.join(detalhes)}" if detalhes else ""
        raise ValueError(
            "Nenhum RMB foi retornado pela API do Bernoulli para as linhas do arquivo. "
            "Verifique se a API esta acessivel pelo backend e se a autenticacao "
            "Bernoulli foi configurada corretamente no .env, seja por "
            "`BERNOULLI_COOKIE`/`BERNOULLI_AUTHORIZATION` ou pelo fluxo de login "
            "automatico (`BERNOULLI_LOGIN_*`)."
            f"{detalhe_texto}"
        )

    resultado = {
        "caminho_saida": caminho_saida,
        "linhas": int(len(saida)),
        "questoes": TOTAL_QUESTOES_BERNOULLI,
        "questoes_encontradas_no_keepedu": int(len(colunas_questoes)),
        "primeira_questao": primeiro_numero_questao,
        "ultima_questao": ultimo_numero_questao,
        "dia": dia,
        "sem_rmb": total_sem_rmb,
        "origem_rmb": "api_bernoulli",
        "rmb_encontrados": int(diagnostico_rmb.get("encontrados", 0)),
    }
    _emitir_progresso(
        progresso_callback,
        99,
        "Arquivo final gerado. Preparando liberacao para download..."
    )
    return resultado
